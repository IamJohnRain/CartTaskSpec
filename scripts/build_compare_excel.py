#!/usr/bin/env python3
"""为每个数据集 case 目录生成「原始 vs 修复」对比 Excel。

对数据集根目录下每个子目录视为一个 case，扫描固定文件约定，导出 6 列对比表：
  1. case_id       — 目录名
  2. 用户指令      — query.txt 内容
  3. 原始 DSL      — card.dsl.jsonl（3 行 JSONL 拼接）
  4. 原始卡片      — card.dsl.png 缩略图
  5. 修复后 DSL    — fix.card.dsl.jsonl（3 行 JSONL 拼接）
  6. 修复后卡片    — fix.card.dsl.png 缩略图

缺失文件会显示 "(缺失)" 而不是抛错，方便快速看出哪些 case 还没修。

示例：
  # 默认：tmp/datasets -> docs/fix_dsl_compared.xlsx
  python scripts/build_compare_excel.py

  # 指定数据集/输出
  python scripts/build_compare_excel.py -d tmp/datasets -o docs/fix_dsl_compared.xlsx

  # 按子串过滤 case + 限制数量
  python scripts/build_compare_excel.py -c Case-19 -n 50
"""
from __future__ import annotations

import argparse
import io
import logging
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError:  # pragma: no cover
    Workbook = None

try:
    from PIL import Image as PILImage
except ImportError:  # pragma: no cover
    PILImage = None

# --------------------------------------------------------------------------- #
# 文件名约定（每个 case 目录）
# --------------------------------------------------------------------------- #
QUERY_NAME = "query.txt"
ORIG_DSL_NAME = "card.dsl.jsonl"
ORIG_PNG_NAME = "card.dsl.png"
FIX_DSL_NAME = "fix.card.dsl.jsonl"
FIX_PNG_NAME = "fix.card.dsl.png"

# 缩略图参数：像素上限（清晰度来源） + Excel 显示宽度（单元格占用）
THUMB_PIXEL_MAX = (600, 1100)
DISPLAY_WIDTH_PX = 240  # 单图列宽；同时放两列，240 兼顾清晰度与表格宽度

log = logging.getLogger("compare-excel")


# --------------------------------------------------------------------------- #
# 读取工具
# --------------------------------------------------------------------------- #
def _read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8").strip()


def _read_dsl_lines(path: Path) -> list[str]:
    lines: list[str] = []
    for raw in path.read_text(encoding="utf-8").splitlines():
        s = raw.strip()
        if s:
            lines.append(s)
    return lines


def find_case_dirs(dataset: Path) -> list[Path]:
    """列出数据集下的 case 目录（所有非隐藏子目录）。

    比 build_score.py 更宽松：只要是子目录就视为 case，缺失文件以 "(缺失)"
    标记在 Excel 中展示，方便统一汇总「待修复」清单。
    """
    return sorted(d for d in dataset.iterdir()
                  if d.is_dir() and not d.name.startswith("."))


# --------------------------------------------------------------------------- #
# 缩略图
# --------------------------------------------------------------------------- #
def _make_thumb_bytes(png_path: Path, col_idx_1based: int, row_idx_1based: int):
    """读取 PNG，生成 (XLImage, disp_w, disp_h)。无 PIL 或无文件返回 (None, 0, 0)。

    像素分辨率保留到 THUMB_PIXEL_MAX 以保证清晰度，显示尺寸固定按
    DISPLAY_WIDTH_PX 等比缩放；通过 OneCellAnchor + 显式 ext(EMU) 锁定
    显示尺寸，避免 openpyxl 按图片 DPI 重算导致图片过大。
    """
    if PILImage is None or not png_path.is_file():
        return None, 0, 0
    try:
        im = PILImage.open(png_path)
        im = im.convert("RGBA")
        im.thumbnail(THUMB_PIXEL_MAX)
        buf = io.BytesIO()
        im.save(buf, format="PNG")
        buf.seek(0)
        disp_w = DISPLAY_WIDTH_PX
        disp_h = round(disp_w * im.height / im.width)
        img = XLImage(buf)
        from openpyxl.drawing.spreadsheet_drawing import (
            OneCellAnchor, AnchorMarker,
        )
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        _from = AnchorMarker(col=col_idx_1based - 1, colOff=0,
                             row=row_idx_1based - 1, rowOff=0)
        ext = XDRPositiveSize2D(cx=disp_w * 9525, cy=disp_h * 9525)
        anchor = OneCellAnchor(_from=_from, ext=ext)
        img.anchor = anchor
        return img, disp_w, disp_h
    except Exception as e:  # noqa: BLE001
        log.warning(f"{png_path.parent.name}: 嵌图失败: {e}")
        return None, 0, 0


# --------------------------------------------------------------------------- #
# 单 case 写入一行 Excel，并打印进度（与 build_score.py 的 [i/total] 日志风格一致）
# --------------------------------------------------------------------------- #
def process_case_for_excel(ws, case_dir: Path, index: int, total: int) -> str:
    """写入一行 6 列对比数据，返回状态标签: 'ok' 或 'ok_warn'。"""
    name = case_dir.name
    lp = f"[{index}/{total}]"
    row_idx = ws.max_row + 1
    max_disp_h = 0
    warns: list[str] = []

    # 1. case_id
    ws.cell(row=row_idx, column=1, value=name).alignment = WRAP_ALIGN

    # 2. 用户指令
    query_path = case_dir / QUERY_NAME
    if query_path.is_file():
        query_text = _read_text(query_path)
    else:
        query_text = "(缺失)"
        warns.append(QUERY_NAME)
    ws.cell(row=row_idx, column=2, value=query_text).alignment = WRAP_ALIGN

    # 3. 原始 DSL
    orig_dsl_path = case_dir / ORIG_DSL_NAME
    if orig_dsl_path.is_file():
        dsl_text = "\n".join(_read_dsl_lines(orig_dsl_path))
    else:
        dsl_text = "(缺失)"
        warns.append(ORIG_DSL_NAME)
    ws.cell(row=row_idx, column=3, value=dsl_text).alignment = WRAP_ALIGN

    # 4. 原始卡片
    orig_png = case_dir / ORIG_PNG_NAME
    if orig_png.is_file():
        img, _, h_px = _make_thumb_bytes(orig_png, 4, row_idx)
        if img is not None:
            ws.add_image(img)
            max_disp_h = max(max_disp_h, h_px)
        else:
            ws.cell(row=row_idx, column=4,
                    value="(无图)").alignment = CENTER_ALIGN
            warns.append(f"{ORIG_PNG_NAME}(无图)")
    else:
        ws.cell(row=row_idx, column=4,
                value="(缺失)").alignment = CENTER_ALIGN
        warns.append(ORIG_PNG_NAME)

    # 5. 修复后 DSL
    fix_dsl_path = case_dir / FIX_DSL_NAME
    if fix_dsl_path.is_file():
        dsl_text = "\n".join(_read_dsl_lines(fix_dsl_path))
    else:
        dsl_text = "(缺失)"
        warns.append(FIX_DSL_NAME)
    ws.cell(row=row_idx, column=5, value=dsl_text).alignment = WRAP_ALIGN

    # 6. 修复后卡片
    fix_png = case_dir / FIX_PNG_NAME
    if fix_png.is_file():
        img, _, h_px = _make_thumb_bytes(fix_png, 6, row_idx)
        if img is not None:
            ws.add_image(img)
            max_disp_h = max(max_disp_h, h_px)
        else:
            ws.cell(row=row_idx, column=6,
                    value="(无图)").alignment = CENTER_ALIGN
            warns.append(f"{FIX_PNG_NAME}(无图)")
    else:
        ws.cell(row=row_idx, column=6,
                value="(缺失)").alignment = CENTER_ALIGN
        warns.append(FIX_PNG_NAME)

    # 行高按本行最高图（两图同宽时取较高那张；无图走默认 28）
    if max_disp_h > 0:
        ws.row_dimensions[row_idx].height = max(28, max_disp_h / 1.333 + 4)
    else:
        ws.row_dimensions[row_idx].height = 28

    # ---- per-case 日志（与 build_score.py 的 [i/total] START/OK/WARN 风格一致） ----
    if warns:
        log.info(f"{lp} WARN  {name} | 缺 {', '.join(warns)}")
        return "ok_warn"
    log.info(f"{lp} OK    {name}")
    return "ok"


# --------------------------------------------------------------------------- #
# 导出 Excel
# --------------------------------------------------------------------------- #
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")


def export_compare_excel(case_dirs: list[Path], out_path: Path) -> int:
    if Workbook is None:
        print("error: 缺少 openpyxl，请 pip install openpyxl", file=sys.stderr)
        return 1
    if PILImage is None:
        print("error: 缺少 Pillow，请 pip install Pillow", file=sys.stderr)
        return 1

    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "fix dsl compared"

    headers = [
        "case_id", "用户指令",
        "原始 DSL", "原始卡片",
        "修复后 DSL", "修复后卡片",
    ]
    ws.append(headers)

    # 表头样式
    hdr_font = Font(bold=True, color="FFFFFFFF")
    hdr_fill = PatternFill(start_color="FF305496", end_color="FF305496",
                           fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # 列宽（与 headers 一一对应；DSL 是长 JSONL，给到 60）
    col_widths = [32, 38, 60, 30, 60, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    total = len(case_dirs)
    counters: dict[str, int] = {"ok": 0, "ok_warn": 0}
    for i, case_dir in enumerate(case_dirs, 1):
        try:
            status = process_case_for_excel(ws, case_dir, i, total)
        except Exception as e:  # noqa: BLE001
            log.error(f"[{i}/{total}] FAIL  {case_dir.name} | {e!r}")
            status = "fail"
        counters[status] = counters.get(status, 0) + 1

    wb.save(out_path)
    log.info(
        f"汇总: 处理 {total} | 完整 {counters['ok']} "
        f"| 缺件 {counters['ok_warn']} | 失败 {counters.get('fail', 0)} "
        f"| 用时 0.0s"
    )
    print(f"已生成 Excel: {out_path} | 总行 {total} "
          f"| 完整 {counters['ok']} | 缺件 {counters['ok_warn']} "
          f"| 失败 {counters.get('fail', 0)}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="为每个 case 目录生成「原始 vs 修复」对比 Excel："
                    "用户 query、原始/修复 DSL 与卡片图片。",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("-d", "--dataset", type=Path, default=Path("tmp/datasets"),
                        help="数据集根目录（其下每个子目录为一个 case）。默认 tmp/datasets。")
    parser.add_argument("-o", "--output", type=Path,
                        default=Path("docs/fix_dsl_compared.xlsx"),
                        help="Excel 输出路径，默认 docs/fix_dsl_compared.xlsx。")
    parser.add_argument("-c", "--case", type=str, default=None,
                        help="只处理名称匹配的 case（子串匹配）。")
    parser.add_argument("-n", "--limit", type=int, default=None,
                        help="最多处理的 case 数量（调试用）。")
    parser.epilog = (
        "示例:\n"
        "  # 默认: tmp/datasets -> docs/fix_dsl_compared.xlsx\n"
        "  python scripts/build_compare_excel.py\n\n"
        "  # 指定数据集与输出\n"
        "  python scripts/build_compare_excel.py -d tmp/datasets "
        "-o docs/fix_dsl_compared.xlsx\n\n"
        "  # 按子串过滤 case + 限制数量\n"
        "  python scripts/build_compare_excel.py -c Case-19 -n 50"
    )
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s | %(levelname)-7s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    if not args.dataset.is_dir():
        print(f"error: 数据集目录不存在: {args.dataset}", file=sys.stderr)
        return 1

    case_dirs = find_case_dirs(args.dataset)
    if args.case:
        case_dirs = [d for d in case_dirs if args.case in d.name]
    if args.limit is not None:
        case_dirs = case_dirs[: args.limit]

    if not case_dirs:
        print("未找到任何符合条件的 case 目录。", file=sys.stderr)
        return 1

    log.info(f"数据集: {args.dataset} | 待处理 case: {len(case_dirs)} 个")
    log.info(f"输出: {args.output}")

    return export_compare_excel(case_dirs, args.output)


if __name__ == "__main__":
    raise SystemExit(main())
