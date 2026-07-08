#!/usr/bin/env python3
"""为每个数据集 case 目录生成 VLM 评分请求体 (score.request.json)，
默认并并发调用 MiniMax chat/completions 评分 API（凭据从 .env 读取）。

评分维度（加权总分 0-10）：
  - 指令遵从 InstructionAdherence   权重 0.4
  - 样式美观 Aesthetic               权重 0.2
  - 排版布局 Layout（堆叠/文字截断） 权重 0.4

请求体顶层为 {"messages":[system, user(多模态)], "response_format": json_object}。
不写 model/temperature 等模型参数，调用 API 时由脚本注入 .env 中的 MINIMAX_MODEL。

示例：
  # 默认：生成请求体 + 调用 API 评分 + 提取改进 DSL 到 -vN 目录
  python scripts/build_score_requests.py -d datasets/cases-600-mix
  python scripts/build_score_requests.py -d datasets/cases-600-mix -c Case-007-01-low-power-007
  python scripts/build_score_requests.py -d datasets/cases-600-mix -n 5 -j 5

  # 只生成请求体，不调用 API
  python scripts/build_score_requests.py -d datasets/cases-600-mix --disable-call-api

  # 全量评分，并发 10，覆盖已有结果
  python scripts/build_score_requests.py -d datasets/cases-600-mix -j 10 --overwrite

  # 调试单个 case（单线程，便于看清日志）
  python scripts/build_score_requests.py -d datasets/cases-600-mix -c Case-007-01-low-power-007 -j 1
"""
from __future__ import annotations

import argparse
import base64
import io
import json
import logging
import os
import re
import shutil
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

try:
    import requests
except ImportError:  # pragma: no cover
    requests = None

log = logging.getLogger("score")

# Excel 导出依赖（仅 --excel 模式用到，缺失时给明确提示）
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except ImportError:  # pragma: no cover
    Workbook = None

try:
    from PIL import Image as PILImage
except ImportError:  # pragma: no cover
    PILImage = None

# --------------------------------------------------------------------------- #
# 文件名约定（每个 case 目录）
# --------------------------------------------------------------------------- #
QUERY_NAME = "query.txt"
TASKSPEC_NAME = "task.taskSpec.json"
DSL_NAME = "card.dsl.jsonl"
PNG_NAME = "card.dsl.png"
OUTPUT_NAME = "score.request.json"
SCORE_RESULT_NAME = "score.result.json"
CARDSPEC_NAME = "card.cardspec.json"

CASE_REQUIRED_FILES = (QUERY_NAME, TASKSPEC_NAME, DSL_NAME, PNG_NAME)

# 改进版 case 目录需要从原 case 复制过来的伴随文件
VERSION_COPY_FILES = (CARDSPEC_NAME, QUERY_NAME, TASKSPEC_NAME)

# --------------------------------------------------------------------------- #
# 评分 system prompt
# --------------------------------------------------------------------------- #
SYSTEM_PROMPT = """你是 HarmonyOS A2UI Form 服务卡片的资深评审专家。你将收到一个 case 的：用户原始指令、TaskSpec 契约、模型生成的 3 行 genui DSL（card.dsl.jsonl）、以及该 DSL 的实际渲染截图。你的任务是对该卡片进行三维加权打分（0-10）、给出中文打分理由，并产出一份改进后的 3 行 DSL。

你的评审依据是协议 .agents/skills/harmony-card-generation-datamodel-first 与 TaskSpec 协议。下文给出关键约束，必须严格据此判断。

=== 评分核心原则（最高优先级，凌驾于一切之上） ===
1. 所有分数必须以渲染截图（card.dsl.png）为唯一且最终的依据。
2. 你必须先认真看图、根据图中实际呈现的效果打分；DSL 与下面的布局公式仅用于"协议合规性"检查与"改进版 DSL"的撰写，绝不能仅凭 DSL 文本/数值就给分。
3. 【排版布局 D3】必须从截图中观察并描述：是否有元素堆叠重叠、文字被裁切/截断（尤其标题、状态、CTA、主数值）、组件溢出卡片边界、Stack 层叠遮挡关键内容。看不到截图中的问题就等于没有问题；反之，只要截图里能看出堆叠/截断，就必须扣分。
4. 【样式美观 D2】必须从截图中判断：配色是否和谐、前景文字在背景上是否清晰可读、留白密度是否舒适、对齐是否统一。不要脱离截图空谈 DSL 里的颜色值或字号。
5. 【指令遵从 D1】参考截图确认信息是否真实、完整地呈现给了用户（而不是只在 DSL 里写了却渲染不出来或被遮挡）。
6. 简言之："截图里看到什么，就评什么。" DSL 只回答"协议是否合法"和"如何改进"。

=== 一、尺寸与 root shell（硬约束） ===
- size 只能是 2x2 或 2x4。
- 2x2：createSurface/root width:140、height:140，root borderRadius:18，clip:true，默认 padding:12，内容区约 116x116。
- 2x4：createSurface/root width:300、height:140，root borderRadius:22，clip:true，默认 padding:12，内容区约 276x116。
- createSurface.catalogId 必须是 "ohos.a2ui.extended.catalog"；每行 version 必须是 "v0.9"。
- updateDataModel.path 必须是 "/"，value 必须原样等于 TaskSpec.dataModel.value。

=== 二、组件与协议范围（硬约束） ===
- 只允许 Text、Image、Divider、Progress、Button、Checkbox、Row、Column、List、Stack。
- 用 Text.content / Image.src / Button.label；禁止 Text.text / Image.url / Button.action。
- 禁止 TextInput、Toggle、Radio、CheckboxGroup、Select、NavContainer、Tabs、TabContent、Web、Grid、If、theme、非 onClick 事件、网络图、SVG、emoji、占位图。
- components 是扁平组件数组；children 只能是组件 id 数组（模板循环才允许 {"componentId":"...","path":"/..."}）。

=== 三、数据绑定、事件、素材（硬约束） ===
- 展示值优先用完整表达式 {{ ... }} 读取 DataModel，路径必须能从 TaskSpec.dataModel.value 推导。
- onClick[].call 与 onClick[].args 必须原样来自 TaskSpec.eventCandidates，不得发明新事件或新参数。
- Image.src / backgroundImage 只能来自 TaskSpec.assetCandidates 或用户明确提供的本地/资源路径。

=== 四、布局诊断参考（辅助交叉验证，不替代看图） ===
注意：本节公式仅用于在截图里发现疑似问题后做交叉验证，或用于撰写改进版 DSL；打分必须以截图为准。当截图与公式结论冲突时，以截图为准。
对每个 Row/Column，可按下式计算并比较：
  - Row 横向占用 = 父 padding.left + 父 padding.right + itemMargin * (子项数 - 1) + Σ(子项 width + 子项 margin.left + 子项 margin.right)
  - Column 纵向占用 = 父 padding.top + 父 padding.bottom + itemMargin * (子项数 - 1) + Σ(子项 height + 子项 margin.top + 子项 margin.bottom)
  - 当「占用 > 父容器 width/height」时判定为潜在溢出（需回到截图确认是否真的出现挤压/重叠/裁剪）。
  - 子项无 width/height 时按估算：文本宽度 ≈ Σ 字符宽度，中文 ≈ fontSize，英文/数字 ≈ 0.6*fontSize，空格/标点 ≈ 0.4*fontSize。
  - 受保护文本（标题、状态、CTA、主指标、用户明确要求字段）不得依赖 ellipsis/clip/marquee 隐藏；放不下应缩字、删弱信息或换尺寸。
  - Stack 叠层不得覆盖受保护文本/CTA/点击元素/进度主值；背景在底层，信息在上层。
  - Button/可点击视觉元素热区 ≥ 24vp，理想接近 40vp。
  - 底部动作区外框底边到卡片底边：2x2 通常 8-12vp、不得 >16vp；2x4 通常 10-14vp、不得 >16vp。
  - 字号阶梯：10/12/14/16/18/20/32/40；间距阶梯：4/8/12/16（微调 2/6/10/14）。同一卡片只保留一个最强字重。

=== 五、评分维度（每项 0-10，给分到小数点后 1 位） ===

D1 指令遵从（权重 0.4）：是否满足 userQuery 的明确诉求（务必结合截图确认信息真实呈现）。
  - 信息是否齐全：用户明确点名要显示的字段/对象是否都在截图里可见且完整。
  - 事件是否落实：userQuery 要求的动作是否转换成 onClick，且 call/args 原样来自 eventCandidates。
  - 素材是否正确：assetCandidates 是否按语义正确使用（背景/图标/插图）。
  - 尺寸与协议：size/root shell/catalogId/version/updateDataModel 等硬约束是否满足。
  - 评分基准：10=全部精准落实且无多余；7-8=主要诉求满足、有轻微瑕疵；4-6=有遗漏或轻微违反协议；1-3=关键诉求缺失或协议违规；0=与指令几乎无关。

D2 样式美观（权重 0.2）：视觉质量，必须基于截图判断。
  - 从截图判断配色是否和谐、渐变是否自然，前景文字在背景上是否清晰可读。
  - 从截图判断字号/字重层级是否清晰、是否只有一个最强字重、间距是否舒适。
  - 从截图判断留白密度是否合理（无明显空洞、无过密），对齐是否统一。
  - 评分基准：10=截图观感优秀、克制、专业；7-8=良好；4-6=可用但有明显土气或不统一；1-3=丑陋或不可读；0=无法入目。

D3 排版布局（权重 0.4）：堆叠、溢出、文字截断（核心排查项，必须基于截图）。
  - 必须从截图中逐一观察：是否存在元素堆叠重叠、文字被裁切/截断、组件溢出卡片边界。
  - 截图中若 Stack 层叠遮盖关键文本/CTA/进度值，即扣重分。
  - 公式核算仅用于交叉验证：若截图显示溢出但公式算不出，仍以截图为准扣分。
  - 检查按钮热区、底部锚定、左右边界共享（均以截图观感为准）。
  - 评分基准：10=截图无任何堆叠/截断，布局干净；7-8=基本成立、个别小瑕疵；4-6=截图存在溢出或局部截断；1-3=截图多处堆叠/文字被裁/CTA 不可用；0=布局崩溃。

加权总分 = D1*0.4 + D2*0.2 + D3*0.4（保留 1 位小数）。

=== 六、输出格式（严格 JSON，禁止任何解释性前后缀） ===
只输出一个 JSON 对象，schema 如下（所有 reason/issue 字段用中文）：
{
  "case_id": "字符串，原样回填",
  "scores": {
    "instruction": {"score": 0-10 数字, "weight": 0.4},
    "aesthetic":   {"score": 0-10 数字, "weight": 0.2},
    "layout":      {"score": 0-10 数字, "weight": 0.4}
  },
  "weighted_total": 0-10 数字,
  "reasons": {
    "instruction": "中文，具体说明指令落实情况与扣分点（结合截图）",
    "aesthetic": "中文，必须基于截图描述美观问题与扣分点",
    "layout": "中文，必须基于截图描述观察到的具体视觉问题，例如'截图中底部按钮文字\"开启省电\"底边被裁切'，而非仅引用 DSL 数值"
  },
  "issues": ["中文具体问题清单，逐条列出"],
  "improved_dsl": [
    "第1行 JSON 字符串：createSurface",
    "第2行 JSON 字符串：updateComponents",
    "第3行 JSON 字符串：updateDataModel"
  ]
}
要求：
- improved_dsl 必须恰好 3 个字符串元素，每个是合法 JSON 行，整体是合法 JSONL；不得包含 markdown 代码块标记。
- 若原 DSL 已无问题，improved_dsl 可以与原 DSL 相同，但不得低于原水平。
- 不要输出 JSON 以外的任何内容。"""


# --------------------------------------------------------------------------- #
# 读取各文件
# --------------------------------------------------------------------------- #
def _read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8").strip()


def _read_taskspec_pretty(path: Path) -> str:
    """读取 task.taskSpec.json 并以 2 空格缩进美化输出，保证模型可读。"""
    obj = json.loads(path.read_text(encoding="utf-8"))
    return json.dumps(obj, ensure_ascii=False, indent=2)


def _read_dsl_lines(path: Path) -> list[str]:
    lines: list[str] = []
    for raw in path.read_text(encoding="utf-8").splitlines():
        s = raw.strip()
        if s:
            lines.append(s)
    return lines


def _encode_image(path: Path) -> str:
    raw = path.read_bytes()
    return "data:image/png;base64," + base64.b64encode(raw).decode("ascii")


# --------------------------------------------------------------------------- #
# 构造请求
# --------------------------------------------------------------------------- #
def build_user_text(case_id: str, query: str, taskspec_pretty: str, dsl_lines: list[str]) -> str:
    dsl_block = "\n".join(dsl_lines) if dsl_lines else "(无)"
    return (
        f"请对以下 case 进行三维加权评分。\n\n"
        f"=== case_id ===\n{case_id}\n\n"
        f"=== 用户原始指令 (query.txt) ===\n{query}\n\n"
        f"=== TaskSpec 契约 (task.taskSpec.json) ===\n{taskspec_pretty}\n\n"
        f"=== 待评审 DSL (card.dsl.jsonl，共 {len(dsl_lines)} 行) ===\n{dsl_block}\n\n"
        f"=== 渲染截图 ===\n⚠️ 本评分以随附的 card.dsl.png 渲染截图为唯一且最终的依据。"
        f"请务必先认真看图，根据图中实际呈现的效果打分；"
        f"上面的 DSL 与 system 中的布局公式仅用于协议合规性检查与改进版 DSL 的撰写，"
        f"绝不能仅凭 DSL 文本/数值给分。排版与美观问题以截图为准。\n\n"
        f"输出要求：严格按 system 中规定的 JSON schema 输出单个 JSON 对象，"
        f"reasons 必须基于截图描述具体视觉现象，improved_dsl 为 3 行合法 JSONL。"
    )


def build_request(case_id: str, query: str, taskspec_pretty: str,
                  dsl_lines: list[str], image_data_url: str | None) -> dict:
    user_text = build_user_text(case_id, query, taskspec_pretty, dsl_lines)

    if image_data_url:
        user_content: list[dict] = [
            {"type": "text", "text": user_text},
            {"type": "image_url", "image_url": {"url": image_data_url}},
        ]
    else:
        user_content = user_text

    return {
        "messages": [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": user_content},
        ],
        "response_format": {"type": "json_object"},
    }


# --------------------------------------------------------------------------- #
# 主流程
# --------------------------------------------------------------------------- #
def find_case_dirs(dataset: Path, require_all_files: bool = True) -> list[Path]:
    """列出数据集下的 case 目录。

    require_all_files=True（默认，用于 API/请求体模式）：必须含
      query/taskspec/dsl/png 全部文件才算合法 case。
    require_all_files=False（用于 --excel 汇总模式）：只要目录名以 'Case-'
      开头即视为 case，允许缺 png/score，便于无评分数据集也入表。
    """
    def _is_case(d: Path) -> bool:
        if not d.is_dir():
            return False
        if require_all_files:
            return all((d / f).is_file() for f in CASE_REQUIRED_FILES)
        return d.name.startswith("Case-")

    return sorted(d for d in dataset.iterdir() if _is_case(d))


def process_case(case_dir: Path, output_name: str, no_image: bool, overwrite: bool) -> str:
    out_path = case_dir / output_name
    if out_path.exists() and not overwrite:
        return "skip_exists"

    try:
        query = _read_text(case_dir / QUERY_NAME)
        taskspec_pretty = _read_taskspec_pretty(case_dir / TASKSPEC_NAME)
        dsl_lines = _read_dsl_lines(case_dir / DSL_NAME)
    except Exception as e:  # noqa: BLE001
        return f"skip_read_error:{e!r}"

    image_data_url: str | None = None
    if not no_image:
        try:
            image_data_url = _encode_image(case_dir / PNG_NAME)
        except Exception as e:  # noqa: BLE001
            return f"skip_image_error:{e!r}"

    request = build_request(
        case_id=case_dir.name,
        query=query,
        taskspec_pretty=taskspec_pretty,
        dsl_lines=dsl_lines,
        image_data_url=image_data_url,
    )

    out_path.write_text(
        json.dumps(request, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return "ok"


# --------------------------------------------------------------------------- #
# .env 与 MiniMax API 配置
# --------------------------------------------------------------------------- #
def _parse_env(path: Path) -> dict:
    env: dict[str, str] = {}
    if not path.is_file():
        return env
    for line in path.read_text(encoding="utf-8").splitlines():
        s = line.strip()
        if not s or s.startswith("#") or "=" not in s:
            continue
        k, v = s.split("=", 1)
        env[k.strip()] = v.strip()
    return env


def load_minimax_config(env_path: Path | None = None) -> dict:
    """从仓库根目录 .env 读取 MiniMax 凭据（已存在的系统环境变量优先）。"""
    resolved = env_path or (Path(__file__).resolve().parents[1] / ".env")
    env = _parse_env(resolved)
    url = os.environ.get("MINIMAX_API_URL") or env.get("MINIMAX_API_URL")
    key = os.environ.get("MINIMAX_API_KEY") or env.get("MINIMAX_API_KEY")
    model = (os.environ.get("MINIMAX_MODEL") or env.get("MINIMAX_MODEL")
             or "MiniMax-M3")
    if not url or not key:
        raise RuntimeError(
            "未配置 MINIMAX_API_URL / MINIMAX_API_KEY，请检查 .env"
        )
    return {"url": url, "key": key, "model": model}


# --------------------------------------------------------------------------- #
# 响应解析（剥离 <think>...、代码块围栏、抽取首个 JSON 对象）
# --------------------------------------------------------------------------- #
def strip_think_and_parse_json(content: str) -> dict:
    if not isinstance(content, str):
        content = str(content)
    content = re.sub(r"<think>.*?</think>", "", content, flags=re.DOTALL)
    content = re.sub(r"<think>.*", "", content, flags=re.DOTALL)
    content = re.sub(r"```(?:json)?", "", content)
    content = content.strip()
    start = content.find("{")
    end = content.rfind("}")
    if start == -1 or end == -1 or end < start:
        raise ValueError("响应中未找到可解析的 JSON 对象")
    return json.loads(content[start : end + 1])


# --------------------------------------------------------------------------- #
# 调用 MiniMax chat/completions
# --------------------------------------------------------------------------- #
def call_minimax(cfg: dict, body: dict, timeout: int, retries: int):
    """流式调用 MiniMax chat/completions，返回 (content, elapsed_sec)。

    M3 是推理模型，非流式请求会在生成完成前整体挂起、极易触发读超时；
    改用 stream=True 逐块读取，块间持续保活，timeout 视为「块间最大间隔」。
    """
    if requests is None:
        raise RuntimeError("缺少 requests 库，请 pip install requests")
    payload = dict(body)
    payload["model"] = cfg["model"]
    payload["stream"] = True
    headers = {
        "Authorization": f"Bearer {cfg['key']}",
        "Content-Type": "application/json",
    }
    last_err: str | None = None
    for attempt in range(retries):
        t0 = time.time()
        chunks: list[str] = []
        try:
            resp = requests.post(
                cfg["url"], json=payload, headers=headers,
                timeout=(15, timeout), stream=True,
            )
        except Exception as e:  # noqa: BLE001
            last_err = f"network(connect): {e!r}"
            log.warning(f"连接异常，{2 ** (attempt + 1)}s 后重试: {e!r}")
            time.sleep(2 ** (attempt + 1))
            continue
        # 可重试状态码
        if resp.status_code == 429 or 500 <= resp.status_code < 600:
            last_err = f"http {resp.status_code}: {resp.text[:200]}"
            log.warning(f"HTTP {resp.status_code}，{2 ** (attempt + 1)}s 后重试")
            time.sleep(2 ** (attempt + 1))
            continue
        if resp.status_code != 200:
            raise RuntimeError(f"http {resp.status_code}: {resp.text[:300]}")
        # 逐块读取 SSE
        try:
            for raw in resp.iter_lines(decode_unicode=True):
                if not raw:
                    continue
                line = raw.strip() if isinstance(raw, str) else raw.decode("utf-8", "ignore").strip()
                if not line.startswith("data:"):
                    continue
                data_str = line[5:].strip()
                if data_str == "[DONE]":
                    break
                try:
                    obj = json.loads(data_str)
                except Exception:  # noqa: BLE001
                    continue
                br = obj.get("base_resp") or {}
                if br.get("status_code", 0) != 0:
                    raise RuntimeError(f"stream base_resp: {br}")
                if "error" in obj:
                    raise RuntimeError(f"stream error: {str(obj['error'])[:300]}")
                ch = obj.get("choices") or []
                if not ch:
                    continue
                delta = ch[0].get("delta", {}) or {}
                c = delta.get("content") or ""
                if not c and delta.get("reasoning_content"):
                    c = delta["reasoning_content"]
                if c:
                    chunks.append(c)
        except requests.exceptions.ConnectionError as e:
            last_err = f"stream(conn): {e!r}"
            log.warning(f"流式中断，{2 ** (attempt + 1)}s 后重试 (已收 {len(chunks)} 块)")
            time.sleep(2 ** (attempt + 1))
            continue
        except requests.exceptions.ReadTimeout as e:
            last_err = f"stream(read-timeout): {e!r}"
            log.warning(f"流式块间超时 {timeout}s，重试 (已收 {len(chunks)} 块)")
            time.sleep(2 ** (attempt + 1))
            continue
        except Exception as e:  # noqa: BLE001
            raise RuntimeError(f"流式读取异常: {e!r}")
        elapsed = time.time() - t0
        content = "".join(chunks).strip()
        if not content:
            last_err = "流式响应 content 为空"
            log.warning("流式响应为空，重试")
            time.sleep(2 ** (attempt + 1))
            continue
        return content, elapsed
    raise RuntimeError(f"API 调用失败（已重试 {retries} 次）: {last_err}")


# --------------------------------------------------------------------------- #
# improved_dsl 归一化 + 版本目录
# --------------------------------------------------------------------------- #
def _normalize_improved_dsl(improved) -> list[str]:
    """把 improved_dsl 归一化为 3 行合法 JSON 字符串；非法则抛异常。"""
    if isinstance(improved, str):
        lines = [l.strip() for l in improved.splitlines() if l.strip()]
    elif isinstance(improved, list):
        lines: list[str] = []
        for item in improved:
            if isinstance(item, str):
                sub = [l.strip() for l in item.splitlines() if l.strip()]
                lines.extend(sub if len(sub) > 1 else [item.strip()])
            else:
                lines.append(json.dumps(item, ensure_ascii=False))
    else:
        raise ValueError("improved_dsl 既非字符串也非数组")
    if len(lines) != 3:
        raise ValueError(f"improved_dsl 应为 3 行，实际 {len(lines)} 行")
    parsed = []
    for ln in lines:
        parsed.append(json.loads(ln))
    if "updateComponents" not in parsed[1]:
        raise ValueError("第 2 行缺少 updateComponents")
    return lines


def next_version_dir(case_dir: Path, version_root: Path) -> Path:
    """在版本输出根目录下查找 {case_name}-vN 的最大版本号，返回 v(N+1) 目录。"""
    name = case_dir.name
    pattern = re.compile(rf"^{re.escape(name)}-v(\d+)$")
    max_v = 0
    for d in version_root.iterdir():
        if d.is_dir():
            m = pattern.match(d.name)
            if m:
                max_v = max(max_v, int(m.group(1)))
    return version_root / f"{name}-v{max_v + 1}"


def create_version_dir(case_dir: Path, version_root: Path,
                       result: dict) -> str:
    """把 improved_dsl 写入新版本目录并复制伴随文件；返回版本目录名。"""
    lines = _normalize_improved_dsl(result.get("improved_dsl"))
    target = next_version_dir(case_dir, version_root)
    target.mkdir(parents=True, exist_ok=False)
    (target / DSL_NAME).write_text("\n".join(lines) + "\n", encoding="utf-8")
    for fn in VERSION_COPY_FILES:
        src = case_dir / fn
        if src.is_file():
            shutil.copy2(src, target / fn)
    return target.name


# --------------------------------------------------------------------------- #
# 单 case 的 API 评分流水（带重试：网络 + 响应解析 + improved_dsl 校验）
# --------------------------------------------------------------------------- #
def _classify_err(e: Exception) -> str:
    """把异常归类成简短中文标签，便于 RETRY 日志分类统计。"""
    msg = str(e)
    if "网络" in msg or "network" in msg.lower() or "http " in msg.lower():
        return "网络"
    if "未找到可解析的 JSON" in msg or "JSONDecodeError" in repr(type(e)):
        return "响应解析"
    if "improved_dsl" in msg or "应为 3 行" in msg or "updateComponents" in msg:
        return "improved_dsl非法"
    return "其它"


def process_case_call_api(case_dir: Path, version_root: Path, cfg: dict,
                          index: int, total: int, overwrite: bool,
                          retries: int, timeout: int) -> str:
    name = case_dir.name
    lp = f"[{index}/{total}]"
    score_path = case_dir / SCORE_RESULT_NAME
    if score_path.exists() and not overwrite:
        log.info(f"{lp} SKIP  {name} (score.result.json 已存在)")
        return "skip"
    body_path = case_dir / OUTPUT_NAME
    if not body_path.is_file():
        try:
            process_case(case_dir, OUTPUT_NAME, no_image=False, overwrite=True)
        except Exception as e:  # noqa: BLE001
            log.error(f"{lp} FAIL  {name} | 构建请求体失败: {e!r}")
            return "fail"
    body = json.loads(body_path.read_text(encoding="utf-8"))
    log.info(f"{lp} START {name}")

    # ---- 外层重试：覆盖 网络 / 响应解析 / improved_dsl非法 三类失败 ----
    result: dict | None = None
    elapsed = 0.0
    dsl_ok = False
    dsl_lines: list[str] | None = None
    last_err: Exception | None = None
    for attempt in range(1, retries + 1):
        try:
            # 传 retries=1 给网络层，避免 × 倍数爆炸；总尝试上限 = 本函数 retries
            content, elapsed = call_minimax(cfg, body, timeout, retries=1)
            result = strip_think_and_parse_json(content)
            # 预校验 improved_dsl（不写盘），合法才视为成功
            dsl_lines = _normalize_improved_dsl(result.get("improved_dsl"))
            dsl_ok = True
            last_err = None
            break
        except Exception as e:  # noqa: BLE001
            last_err = e
            kind = _classify_err(e)
            if attempt < retries:
                backoff = 2 ** attempt
                log.info(
                    f"{lp} RETRY {name} | {attempt}/{retries}次({kind}): "
                    f"{e} | {backoff}s后重试"
                )
                time.sleep(backoff)
            else:
                log.warning(
                    f"{lp} RETRY {name} | {attempt}/{retries}次({kind})"
                    f" 已用尽: {e}"
                )

    # ---- 降级判断 ----
    # 1) result 从未解析成功 → 彻底失败（不落盘）
    if result is None:
        log.error(f"{lp} FAIL  {name} | 重试{retries}次仍失败: {last_err}")
        return "fail"

    # 2) 保存评分（即使 dsl 非法也要留下分数）
    score_path.write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    wt = result.get("weighted_total")
    wt_str = f"{wt}" if isinstance(wt, (int, float)) else "?"

    # 3) improved_dsl 合法 → 写版本目录；非法 → WARN 但保留分数
    vtag = ""
    if dsl_ok and dsl_lines is not None:
        try:
            vtag = create_version_dir(case_dir, version_root, result)
        except Exception as e:  # noqa: BLE001
            log.warning(f"{lp} WARN  {name} | 版本目录创建失败: {e}")
    else:
        log.warning(
            f"{lp} WARN  {name} | improved_dsl 重试{retries}次仍非法，"
            f"已保留分数但不生成版本目录"
        )
    log.info(
        f"{lp} OK    {name} | {elapsed:.1f}s | total={wt_str} "
        f"| -> {vtag or '(无版本目录)'}"
    )
    return "ok"


def run_api_batch(case_dirs: list[Path], args: argparse.Namespace) -> int:
    cfg = load_minimax_config(args.env)
    version_root = args.version_out or args.dataset
    version_root.mkdir(parents=True, exist_ok=True)
    log.info(
        f"API: {cfg['url']} | model={cfg['model']} "
        f"| concurrency={args.concurrency} | retries={args.retries} "
        f"| timeout={args.timeout}s"
    )
    log.info(f"版本输出目录: {version_root}")
    log.info(f"待处理 case: {len(case_dirs)} 个")

    total = len(case_dirs)
    t0 = time.time()
    counters = {"ok": 0, "fail": 0, "skip": 0}

    def task(i: int, d: Path) -> str:
        try:
            return process_case_call_api(
                case_dir=d,
                version_root=version_root,
                cfg=cfg,
                index=i + 1,
                total=total,
                overwrite=args.overwrite,
                retries=args.retries,
                timeout=args.timeout,
            )
        except Exception as e:  # noqa: BLE001
            log.error(f"[{i + 1}/{total}] FAIL  {d.name} | 未捕获异常: {e!r}")
            return "fail"

    with ThreadPoolExecutor(max_workers=args.concurrency) as ex:
        futs = {ex.submit(task, i, d): d.name for i, d in enumerate(case_dirs)}
        for fut in as_completed(futs):
            st = fut.result()
            counters[st] = counters.get(st, 0) + 1

    elapsed = time.time() - t0
    log.info(
        f"汇总: 完成 {counters['ok']} | 失败 {counters['fail']} "
        f"| 跳过 {counters['skip']} | 用时 {elapsed:.1f}s"
    )
    return 0 if counters["fail"] == 0 else 1


# --------------------------------------------------------------------------- #
# Excel 分数汇总（只读 score.result.json，绝不调 API）
# --------------------------------------------------------------------------- #
THUMB_PIXEL_MAX = (600, 1100)  # PNG 真实像素上限（决定清晰度）
DISPLAY_WIDTH_PX = 150         # Excel 内显示宽度 px（决定单元格占用）

# 条件着色：分数低 → 红，中 → 黄，高 → 绿
_FILL_RED = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE",
                        fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFFFEB9C", end_color="FFFFEB9C",
                           fill_type="solid")
_FILL_GREEN = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE",
                          fill_type="solid")


def _score_fill(score) -> PatternFill | None:
    if not isinstance(score, (int, float)):
        return None
    if score < 4:
        return _FILL_RED
    if score < 7:
        return _FILL_YELLOW
    return _FILL_GREEN


def _load_score(case_dir: Path) -> dict | None:
    p = case_dir / SCORE_RESULT_NAME
    if not p.is_file():
        return None
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:  # noqa: BLE001
        log.warning(f"跳过 {case_dir.name}: score.result.json 解析失败: {e}")
        return None


def _make_thumb_bytes(png_path: Path, row_idx: int):
    """返回 (XLImage, disp_w, disp_h)；无 PNG 或无 PIL 返回 (None, 0, 0)。

    只保留原图上 1/3（长截图的最上方卡片区域）。
    像素分辨率与 Excel 显示尺寸解耦：PNG 保留高像素(≤THUMB_PIXEL_MAX)
    以保证清晰度，显示尺寸固定按 DISPLAY_WIDTH_PX 等比缩放。
    通过 OneCellAnchor + 显式 ext(EMU) 强制显示尺寸，避免 openpyxl 按
    图片 DPI 重算导致图片被画得过大。
    """
    if PILImage is None or not png_path.is_file():
        return None, 0, 0
    try:
        im = PILImage.open(png_path)
        im = im.convert("RGBA")
        # 先裁剪：从顶部保留 1/3 高度
        im = im.crop((0, 0, im.width, max(1, im.height // 3)))
        # 像素分辨率：保留到 THUMB_PIXEL_MAX（清晰度来源）
        im.thumbnail(THUMB_PIXEL_MAX)
        buf = io.BytesIO()
        im.save(buf, format="PNG")
        buf.seek(0)
        # 显示尺寸：与像素解耦，固定显示宽度等比缩放
        disp_w = DISPLAY_WIDTH_PX
        disp_h = round(disp_w * im.height / im.width)
        img = XLImage(buf)
        # 用 OneCellAnchor 显式指定锚点(C 列 = col index 2)与显示尺寸(EMU)。
        # 1 px @ 96 DPI = 9525 EMU
        from openpyxl.drawing.spreadsheet_drawing import (
            OneCellAnchor, AnchorMarker,
        )
        from openpyxl.drawing.xdr import XDRPositiveSize2D
        _from = AnchorMarker(col=2, colOff=0,
                             row=row_idx - 1, rowOff=0)  # row 0-based
        ext = XDRPositiveSize2D(cx=disp_w * 9525, cy=disp_h * 9525)
        anchor = OneCellAnchor(_from=_from, ext=ext)
        img.anchor = anchor
        return img, disp_w, disp_h
    except Exception as e:  # noqa: BLE001
        log.warning(f"{png_path.parent.name}: 嵌图失败: {e}")
        return None, 0, 0
    except Exception as e:  # noqa: BLE001
        log.warning(f"{png_path.parent.name}: 嵌图失败: {e}")
        return None, 0, 0


def export_excel(case_dirs: list[Path], dataset: Path,
                 out_path: Path | None) -> int:
    if Workbook is None:
        print("error: 缺少 openpyxl，请 pip install openpyxl", file=sys.stderr)
        return 1
    if PILImage is None:
        print("error: 缺少 Pillow，请 pip install Pillow", file=sys.stderr)
        return 1

    if out_path is None or str(out_path) == "":
        out_path = dataset / "score_summary.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "score summary"

    headers = [
        "case_id", "用户指令", "渲染图", "加权总分",
        "指令遵从(0.4)", "样式美观(0.2)", "排版布局(0.4)",
        "问题数", "问题清单", "排版原因", "指令原因", "美观原因",
    ]
    ws.append(headers)
    # 表头样式
    hdr_font = Font(bold=True, color="FFFFFFFF")
    hdr_fill = PatternFill(start_color="FF305496", end_color="FF305496",
                           fill_type="solid")
    for col_idx in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    wrap_align = Alignment(wrap_text=True, vertical="top")
    center_align = Alignment(horizontal="center", vertical="center")
    # 列宽（与 headers 一一对应：用户指令列较宽）
    col_widths = [32, 50, 24, 11, 13, 13, 13, 8, 50, 50, 45, 45]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    written = 0
    no_score = 0
    for case_dir in case_dirs:
        result = _load_score(case_dir)
        has_score = result is not None
        if not has_score:
            no_score += 1
        scores = (result.get("scores", {}) if has_score else {}) or {}
        ins = scores.get("instruction", {}) or {}
        aes = scores.get("aesthetic", {}) or {}
        lay = scores.get("layout", {}) or {}
        wt = result.get("weighted_total") if has_score else None
        issues = (result.get("issues", []) if has_score else []) or []
        reasons = (result.get("reasons", {}) if has_score else {}) or {}
        row_idx = ws.max_row + 1

        # 读取用户原始 query（容错缺失）
        query_path = case_dir / QUERY_NAME
        query_text = (query_path.read_text(encoding="utf-8").strip()
                      if query_path.is_file() else "")

        ws.cell(row=row_idx, column=1, value=case_dir.name).alignment = wrap_align
        ws.cell(row=row_idx, column=2, value=query_text).alignment = wrap_align
        ws.cell(row=row_idx, column=4, value=wt).alignment = center_align
        ws.cell(row=row_idx, column=5,
                value=ins.get("score")).alignment = center_align
        ws.cell(row=row_idx, column=6,
                value=aes.get("score")).alignment = center_align
        ws.cell(row=row_idx, column=7,
                value=lay.get("score")).alignment = center_align
        ws.cell(row=row_idx, column=8,
                value=len(issues) if has_score else None).alignment = center_align
        ws.cell(row=row_idx, column=9,
                value=("\n".join(f"· {x}" for x in issues)
                       if has_score else None)).alignment = wrap_align
        ws.cell(row=row_idx, column=10,
                value=(reasons.get("layout") or None)).alignment = wrap_align
        ws.cell(row=row_idx, column=11,
                value=(reasons.get("instruction") or None)).alignment = wrap_align
        ws.cell(row=row_idx, column=12,
                value=(reasons.get("aesthetic") or None)).alignment = wrap_align

        # 条件着色（总分与三个分维度，列号随插入用户指令列后 +1）
        for col, val in ((4, wt), (5, ins.get("score")),
                         (6, aes.get("score")), (7, lay.get("score"))):
            fill = _score_fill(val)
            if fill:
                ws.cell(row=row_idx, column=col).fill = fill

        # 嵌入缩略图（C 列，显示尺寸由 _make_thumb_bytes 内的 OneCellAnchor 决定）
        img, w_px, h_px = _make_thumb_bytes(case_dir / PNG_NAME, row_idx)
        if img is not None:
            ws.add_image(img)  # anchor 已在 _make_thumb_bytes 内预设
            # 行高按缩略图显示高度（1pt ≈ 1.333px）
            ws.row_dimensions[row_idx].height = max(28, h_px / 1.333 + 4)
        else:
            ws.cell(row=row_idx, column=3, value="(无图)").alignment = center_align
            ws.row_dimensions[row_idx].height = 28

        written += 1

    wb.save(out_path)
    print(f"已生成 Excel: {out_path} | 总行 {written} "
          f"| 其中无评分(留空) {no_score}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description="为每个 case 目录生成 VLM 评分请求体 requestbody.json，"
                    "默认并发调用 MiniMax 评分 API。"
    )
    parser.add_argument("-d", "--dataset", type=Path, default=Path("datasets/cases-600-mix"),
                        help="数据集根目录（其下每个子目录为一个 case）。")
    parser.add_argument("-c", "--case", type=str, default=None,
                        help="只处理名称匹配的 case（子串匹配）。")
    parser.add_argument("-n", "--limit", type=int, default=None,
                        help="最多处理的 case 数量（调试用）。")
    parser.add_argument("-o", "--output-name", type=str, default=OUTPUT_NAME,
                        help=f"请求体输出文件名，默认 {OUTPUT_NAME}。")
    parser.add_argument("--no-image", action="store_true",
                        help="不编码 PNG，生成纯文本请求体。")
    parser.add_argument("--overwrite", action="store_true",
                        help="覆盖已存在的输出/评分结果。")
    parser.add_argument("--disable-call-api", action="store_true",
                        help="只生成请求体，不调用 MiniMax 评分 API（默认会调用）。")
    parser.add_argument("-j", "--concurrency", type=int, default=5,
                        help="调用 API 时的并发 case 数，默认 5。")
    parser.add_argument("-r", "--retries", type=int, default=3,
                        help="单 case 失败（网络/响应解析/improved_dsl非法）的"
                             "总重试次数，默认 3。")
    parser.add_argument("-t", "--timeout", type=int, default=300,
                        help="单次 API 请求超时秒数（流式模式下为块间最大"
                             "间隔），默认 300（5 分钟）。")
    parser.add_argument("-e", "--env", type=Path, default=None,
                        help=".env 文件路径，默认仓库根目录下的 .env。")
    parser.add_argument("-v", "--version-out", type=Path, default=None,
                        help="改进版 -vN case 的输出根目录；不传则默认与"
                             "数据集同目录（-v 指定的目录不存在会自动创建）。")
    parser.add_argument("-x", "--excel", type=Path, nargs="?", const=None,
                        default=None,
                        help="只读现有 score.result.json 汇总成分数 Excel"
                             "（不调 API）。可跟输出路径；省略则输出到"
                             " <dataset>/score_summary.xlsx。")
    parser.epilog = (
        "示例:\n"
        "  # 默认:生成请求体 + 调用 API 评分 + 提取改进 DSL 到 -vN 目录\n"
        "  python scripts/build_score.py -d datasets/cases-600-mix\n"
        "  python scripts/build_score.py -d datasets/cases-600-mix "
        "-c Case-007-01-low-power-007\n"
        "  python scripts/build_score.py -d datasets/cases-600-mix "
        "-n 5 -j 5\n\n"
        "  # 把改进版 case 输出到独立目录(不污染原数据集,目录不存在自动创建)\n"
        "  python scripts/build_score.py -d datasets/cases-600-mix "
        "-v datasets/cases-600-mix-iter1\n\n"
        "  # 只生成请求体,不调用 API\n"
        "  python scripts/build_score.py -d datasets/cases-600-mix "
        "--disable-call-api\n\n"
        "  # 全量评分,并发 10,覆盖已有结果\n"
        "  python scripts/build_score.py -d datasets/cases-600-mix "
        "-j 10 --overwrite\n\n"
        "  # 调试单 case(单线程)\n"
        "  python scripts/build_score.py -d datasets/cases-600-mix "
        "-c Case-007-01-low-power-007 -j 1\n\n"
        "  # 只读现有 score.result.json 汇总成 Excel(不调 API)\n"
        "  python scripts/build_score.py -d datasets/cases-600-mix --excel\n"
        "  python scripts/build_score.py -d datasets/cases-600-mix "
        "--excel out/report.xlsx"
    )
    parser.formatter_class = argparse.RawDescriptionHelpFormatter
    args = parser.parse_args()

    # --excel 模式：只读现有评分，绝不调 API，与其它模式互斥
    excel_mode = "--excel" in sys.argv or "-x" in sys.argv
    call_api = not args.disable_call_api and not excel_mode
    if call_api or excel_mode:
        logging.basicConfig(
            level=logging.INFO,
            format="%(asctime)s | %(levelname)-7s | %(message)s",
            datefmt="%Y-%m-%d %H:%M:%S",
        )

    if not args.dataset.is_dir():
        print(f"error: 数据集目录不存在: {args.dataset}", file=sys.stderr)
        return 1

    case_dirs = find_case_dirs(args.dataset, require_all_files=not excel_mode)
    if args.case:
        case_dirs = [d for d in case_dirs if args.case in d.name]
    if args.limit is not None:
        case_dirs = case_dirs[: args.limit]

    if not case_dirs:
        print("未找到任何符合条件的 case 目录。", file=sys.stderr)
        return 1

    if excel_mode:
        return export_excel(case_dirs, args.dataset, args.excel)

    if call_api:
        return run_api_batch(case_dirs, args)

    ok = 0
    skipped = 0
    errors = 0
    for d in case_dirs:
        status = process_case(
            case_dir=d,
            output_name=args.output_name,
            no_image=args.no_image,
            overwrite=args.overwrite,
        )
        if status == "ok":
            ok += 1
            print(f"  [ok]   {d.name}")
        elif status == "skip_exists":
            skipped += 1
            print(f"  [skip] {d.name} (已存在，--overwrite 覆盖)")
        else:
            errors += 1
            print(f"  [err]  {d.name} -> {status}", file=sys.stderr)

    total = len(case_dirs)
    print(f"\n汇总：处理 {total} 个 | 成功 {ok} | 跳过 {skipped} | 错误 {errors}")
    return 0 if errors == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
